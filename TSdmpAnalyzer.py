#alteonToExcel
#Created and maintained by Steve Harris - Steven.Harris@radware.com
import os
from datetime import date
from clsAlteon import *
import openpyxl


#####Adjustable settings#####
config_path = "./TSDmp/"
report_path = "./Reports/"
filename = f'./Reports/AlteonReport.{date.today().strftime("%d %b %Y")}.xlsx'
##########


def get_readme_version(path="Readme.txt"):
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            text = f.read()

        # Regex: find "# Version control" followed by newline, then capture the next line
        match = re.search(r"# Version control\s*\n([^\n]+)", text, re.IGNORECASE)
        if match:
            return match.group(1).strip()
        return None
    except:
        return None

print(f"\nTSdmpAnalyzer version {get_readme_version()}\n\n\
Please note, this is a new script. It has been tested against a small number of files. \n\
It is strongly recommended that you manually review your TSdmp after running the script to make sure nothing was missed.\n\
\n\
If you notice any problems, please contact Steve Harris at Steven.Harris@radware.com\n\
\n")

if not os.path.exists(config_path):
    os.makedirs(config_path)

if not os.path.exists('Reports'):#Todo: Fix to use global variable
    os.makedirs('Reports')

outputRows = []
outputHeaders = clsTSdmp.getRecommendationHeaders()
outputHeaders.extend(clsAlteonConfig.getRecommendationHeaders())
for path, dir, files in os.walk(config_path):
    #Don't process files in the NoProcess subfolder.
    if (path.startswith( f'{config_path}NoProcess')):
        continue

    for file in files:
        
        if file == "DeleteMe. TSdmp files go here":
            pass
        elif file.lower().endswith(".tgz"):
            #File is a TechData file
            #try:
                print("TechData file: " + path + '/' + file)
                techData=clsTechData(path,file)
                outputRows.append(techData.outputCells)
            #except Exception as err:
            #    print(f'Error processing {path + file} {err}')
            #    #outputRows.append([{'text' : file, 'color' : 'FFC7CE'},{'text' : f"Error reading file\n{err}", 'color' : 'FFC7CE'} ])
            #    outputRows.append([{'text' : file, 'color' : 'FFC7CE'},{'text' : f"Error reading file", 'color' : 'FFC7CE'} ])
        elif file.lower().endswith('.txt') or file.count('.') == 0:
            #try:
            with open(path + "/" + file, 'r', encoding='utf8', errors="ignore") as f:
                first_line = f.readline().strip()
                f.seek(0)
                if first_line.count("script start ") == 1 and first_line.endswith("/**** DO NOT EDIT THIS LINE!"):
                    #File is a configuration
                    print(f"Config file: {path}/{file}")
                    alteonConfig = clsAlteonConfig(f.read())
                    outrow =[
                        alteonConfig.getHostname(),
                        {'text':file},
                        alteonConfig.getManagementIP(),
                        None,# BaseMac,
                        None,# LicenseMac,
                        alteonConfig.getModel(),
                        alteonConfig.getVersion(),
                        alteonConfig.getDate(),
                        None,# self.vADCs,
                        None,# self.Uptime,
                        None,# self.HAInfo,
                        None,# self.ApplyFlags,
                        None,# self.SSHSessions,
                        None,# self.AllocationFailures,
                        None,# self.LicenseUtilization,
                        alteonConfig.getSessCap(),
                        None,# self.PanicDumps,
                        None,# self.AlarmingSyslogs,
                        alteonConfig.getConfiguredServices(),
                        alteonConfig.getManagementACLs(),
                        None,# self.RealServerStates,
                        None,# self.VirtualServerStates,
                        None,# self.Fans,
                        None,# self.Temp,
                        None,# self.interfaces,
                        None,# self.PortsEther,
                        None# self.PortsIf,
                    ]
                    outrow.extend(alteonConfig.getConfigRecommendationColumns())
                    outputRows.append(outrow)
                else:
                    #File is a tsdmp file
                    TSdmp = ''
                    print(f"Tsdmp file: {path}/{file}")

                    TSdmp = clsTSdmp(f.read(), f"{path}/{file}")
                    if len(TSdmp.raw) > 0:
                        print("TSdmp found. Analyzing")
                        outputRows.append(TSdmp.analyze())
                        #outputHeaders = TSdmp.getRecommendationHeaders()
                    else:
                        print(f'Error processing {path + "/" + file} - Empty file.')
                        outputRows.append([{'text' : file, 'color' : 'FFC7CE'},{'text' : f"Error: Empty file.", 'color' : 'FFC7CE'} ])
            #except Exception as err:
            #    print(f'Error processing {path + "/" + file} {err}')
            #    outputRows.append([{'text' : file, 'color' : 'FFC7CE'},{'text' : f"Error reading file", 'color' : 'FFC7CE'} ])
    

print("\nParsing Complete.")
########################################################
print("Creating subnet overlap report for all processed tsdmps")
import ipaddress
from collections import defaultdict

subnet_map = defaultdict(list)  # subnet -> list of (device_index, ip)

def parse_lines(text, device_idx):
    lines = text.strip().splitlines()
    for line in lines:
        parts = line.strip().split()
        if not parts:
            continue
        try:
            if ':' in parts[0]:  # IPv6
                ip = parts[0]
                prefix = parts[1] if len(parts) > 1 else '64'
                net = ipaddress.IPv6Network(f"{ip}/{prefix}", strict=False)
            else:  # IPv4
                ip = parts[0]
                if len(parts) > 1:
                    mask_or_prefix = parts[1]
                    if '.' in mask_or_prefix:  # subnet mask
                        net = ipaddress.IPv4Network(f"{ip}/{mask_or_prefix}", strict=False)
                    else:  # CIDR prefix
                        net = ipaddress.IPv4Network(f"{ip}/{mask_or_prefix}", strict=False)
                else:
                    net = ipaddress.IPv4Network(f"{ip}/24", strict=False)

            subnet_map[net].append((device_idx, ip))

        except ValueError as e:
            print(f"Skipping invalid line: '{line}' – {e}")

# Parse Subnet column
for idx, row in enumerate(outputRows):
    if row and row[24]:
        text = row[24].get('text','') 
        if text.strip():
            parse_lines(text, idx)

# Report shared subnets with sorted IPs
subnet_output = ""
for subnet, entries in subnet_map.items():
    if len(entries) > 1:
        subnet_output += f"\nSubnet {subnet} has overlapping IPs:\n"
        # Sort entries by parsed IP address
        sorted_entries = sorted(entries, key=lambda x: ipaddress.ip_address(x[1]))
        for device_idx, ip in sorted_entries:
            subnet_output += f"  {ip} - {outputRows[device_idx][0]['text']}\n"
if subnet_output == "":
    subnet_output = "No overlapping subnets found."
os.makedirs("./Reports", exist_ok=True)
report_path = "./Reports/Subnet Overlap Report.txt"

with open(report_path, "w", encoding="utf-8") as f:
    f.write(subnet_output.strip() + "\n")

########################################################
print("Generating spreadsheet")
wb = openpyxl.Workbook()
sheet = wb.active
# outputHeaders = ["Hostname",
#         "File Name",
#         "Management IP",
#         "Base MAC",
#         "License MAC",
#         "Model",
#         "SW Version",
#         "Date",
#         "VX vADCs",
#         "Time since last reboot",
#         "HA Info",
#         "Apply/Save/Sync",
#         "Stale SSH Entries",
#         "PIP failures",
#         "License \ Limit \ Peak \ Current",
#         "Session Table Setting",
#         "Panic dumps",
#         "ALERT|CRITICAL|WARNING syslog entries (last 200)",
#         "Network Services",
#         "Management ACLs",
#         "Real Servers (Not up)",
#         "Virtual Servers",
#         "Fan state",
#         "Temperature state",
#         "L3 Interfaces",
#         "Ethernet port issues",
#         "Interface issues",
#         "Unused Config Elements"
#         ]

sheet.append(outputHeaders)
for cell in sheet["1:1"]:
    cell.font = openpyxl.styles.Font(bold=True)

#Sort the output:
outputRows.sort(key=lambda row: row[0]['text'])

#Output the data into rows
curRow = 2
for dataRow in outputRows:
    curCol = 1
    for dataCell in dataRow:
        curCell = sheet.cell(row=curRow, column=curCol)
        curCell.alignment = openpyxl.styles.Alignment(wrapText=True, vertical='top')
        curCell.border = openpyxl.styles.borders.Border(
                                                        left = openpyxl.styles.borders.Side(style = 'thin'),
                                                        right = openpyxl.styles.borders.Side(style = 'thin'),
                                                        top = openpyxl.styles.borders.Side(style = 'thick'),
                                                        bottom = openpyxl.styles.borders.Side(style = 'thick')
                                                        )
        if dataCell:
            # an = at the front of a line indicates a formula in excel. Add a space to the front to correct.
            if dataCell['text'] and "=" in dataCell['text']:
                dataCell['text'] = re.sub(r'^=',' =', dataCell['text'])
            #curCell.value = dataCell['text']
            curCell.value = openpyxl.cell.cell.ILLEGAL_CHARACTERS_RE.sub("", dataCell.get('text',''))
            if 'color' in dataCell:
                my_fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=dataCell['color'])
                curCell.fill = my_fill
            
        curCol += 1
    curRow += 1
        
        
# Auto-fit columns to fit the data
for column in sheet.columns:
    max_length = 0
    column_letter = column[0].column_letter
    has_data = False  # Track if this column has any real data

    for cell in column:
        value = cell.value
        if value is not None and str(value).strip() != "":
            if cell.row > 1 and str(value).strip() != "N/A":
                has_data = True
            try:
                wrapped_text = str(value).strip()
                for line in wrapped_text.splitlines():
                    if len(line) > max_length:
                        max_length = len(line)
            except Exception as e:
                print(e)
                pass
    
    adjusted_width = (max_length + 1.5) * 1.02  # Adjust multiplier as needed
    sheet.column_dimensions[column_letter].width = min(adjusted_width, 200)
    if not has_data:
        sheet.column_dimensions[column_letter].hidden = True

#Freeze the header row
sheet.freeze_panes = sheet['B2']

#Save the worksheet

retry=True
while (retry == True):
    try:
        print("Saving to " + filename)
        wb.save(filename)
        print("\nOutput saved successfully!")
        retry = False
    except Exception as e:
        print(f'\nError writing to {filename}\n    {e}')
        print("Press enter to retry. Press any other key to abort")
        input = input()
        if len(input) > 0:
            retry = False
    except:
        print("Write failed")
        retry = False



os._exit(0)


